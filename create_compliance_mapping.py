from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = Workbook()

# ============================================================
# TAB 1: NIST AI RMF
# ============================================================
ws1 = wb.active
ws1.title = "NIST AI RMF"

headers1 = [
    "Project",
    "Control / Activity",
    "NIST AI RMF Function",
    "Evidence Artifact",
    "Security Outcome"
]

data1 = [
    [
        "Project 1",
        "AI platform data-flow and identity trust-boundary analysis",
        "MAP",
        "OWASP Threat Dragon data-flow diagram",
        "Identifies identity boundaries and trust relationships"
    ],
    [
        "Project 1",
        "STRIDE threat identification",
        "MAP",
        "STRIDE threat panel",
        "Identifies RAG poisoning, information disclosure and privilege risks"
    ],
    [
        "Project 1",
        "Risk-ranked threat register",
        "MAP",
        "12-entry threat register",
        "Ranks threats using likelihood and impact"
    ],
    [
        "Project 1",
        "MITRE ATLAS technique mapping",
        "MAP",
        "MITRE ATLAS mapping evidence",
        "Connects threats to AI attack techniques"
    ],
    [
        "Project 2",
        "Auth0 application and API configuration",
        "GOVERN",
        "Auth0 Applications and API evidence",
        "Establishes identity and access controls"
    ],
    [
        "Project 2",
        "RBAC and API scopes",
        "GOVERN",
        "Auth0 role and permission scope evidence",
        "Supports least-privilege authorization"
    ],
    [
        "Project 2",
        "Multi-factor authentication",
        "GOVERN",
        "Auth0 MFA evidence",
        "Strengthens user authentication"
    ],
    [
        "Project 2",
        "Ed25519 agent identity binding",
        "GOVERN",
        "Auth0 Post-Login Action and Ed25519 evidence",
        "Associates authenticated requests with agent identity"
    ],
    [
        "Project 3",
        "Red-team attack testing",
        "MEASURE",
        "Project 3 attack evidence",
        "Measures security weaknesses through adversarial testing"
    ],
    [
        "Project 3",
        "CVSS scoring and risk measurement",
        "MEASURE",
        "CVSS finding table",
        "Quantifies attack severity"
    ],
    [
        "Project 4",
        "Indirect prompt injection guardrail",
        "MANAGE",
        "guardrail_test.py output",
        "Prevents malicious prompt instructions from being processed"
    ],
    [
        "Project 4",
        "JWT credential detection and output redaction",
        "MANAGE",
        "jwt_redaction_test.py output",
        "Prevents credential exposure in responses"
    ],
    [
        "Project 4",
        "Ed25519 message signing and signature verification",
        "MANAGE",
        "generate_keys.py and signature_test.py",
        "Protects agent-message integrity"
    ],
    [
        "Project 4",
        "Refresh-token lifetime and rotation",
        "MANAGE",
        "Auth0 dashboard and refresh-token replay test",
        "Prevents refresh-token replay"
    ],
    [
        "Project 4",
        "API and agent anomaly detection",
        "MEASURE",
        "anomaly_detection.py output",
        "Detects abnormal API and agent behavior"
    ]
]

ws1.append(headers1)
for row in data1:
    ws1.append(row)


# ============================================================
# TAB 2: OWASP LLM TOP 10
# ============================================================
ws2 = wb.create_sheet("OWASP LLM Top 10")

headers2 = [
    "OWASP Category",
    "Project 3 Attack",
    "Project 4 Control",
    "Evidence Artifact"
]

data2 = [
    [
        "LLM01 - Prompt Injection",
        "Indirect prompt injection and instruction override",
        "NeMo input guardrail blocks malicious instructions",
        "guardrail_test.py"
    ],
    [
        "LLM06 - Sensitive Information Disclosure",
        "Credential and JWT extraction attack",
        "JWT detection and output redaction",
        "jwt_redaction_test.py"
    ],
    [
        "LLM07 - Insecure Plugin Design",
        "RAG / MCP poisoning and unauthorized tool manipulation",
        "Input guardrails and anomaly detection",
        "guardrail_test.py and anomaly_detection.py"
    ],
    [
        "LLM09 - Misinformation",
        "Malicious instructions attempting to manipulate agent output",
        "Input validation, guardrails and human review",
        "guardrail_test.py"
    ]
]

ws2.append(headers2)
for row in data2:
    ws2.append(row)


# ============================================================
# TAB 3: COMPLIANCE GAPS
# ============================================================
ws3 = wb.create_sheet("Compliance Gaps")

headers3 = [
    "Finding",
    "Obligation",
    "Risk",
    "Recommendation"
]

data3 = [
    [
        "No centralized secrets management platform implemented",
        "Secrets must be securely stored and access-controlled",
        "API keys or credentials may be exposed in code or files",
        "Implement an approved secrets manager and remove secrets from source code"
    ],
    [
        "No automated continuous compliance monitoring",
        "Security controls should be reviewed and monitored regularly",
        "Control failures may remain undetected",
        "Implement automated control monitoring and periodic compliance reviews"
    ],
    [
        "No formal AI data classification and retention policy",
        "Sensitive data must be identified, protected and retained appropriately",
        "Sensitive data may be overexposed or retained longer than necessary",
        "Create a formal AI data classification, retention and deletion policy"
    ]
]

ws3.append(headers3)
for row in data3:
    ws3.append(row)


# ============================================================
# FORMATTING
# ============================================================
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="B7B7B7")

for ws in [ws1, ws2, ws3]:

    # Header formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    # Cell formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ============================================================
# COLUMN WIDTHS
# ============================================================

# NIST AI RMF - compact widths
ws1.column_dimensions["A"].width = 12
ws1.column_dimensions["B"].width = 38
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 35
ws1.column_dimensions["E"].width = 38

# OWASP LLM Top 10
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 42
ws2.column_dimensions["D"].width = 35

# Compliance Gaps
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 42
ws3.column_dimensions["C"].width = 42
ws3.column_dimensions["D"].width = 50


# ============================================================
# COMPACT ROW HEIGHTS - REDUCES LARGE GAPS
# ============================================================

for ws in [ws1, ws2, ws3]:
    ws.row_dimensions[1].height = 25

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 30


# ============================================================
# SAVE WORKBOOK
# ============================================================

output_file = "Compliance_Mapping.xlsx"
wb.save(output_file)

print("=" * 65)
print("SECURENOVA PROJECT 5 - COMPLIANCE MAPPING")
print("=" * 65)
print()
print("NIST AI RMF tab created successfully.")
print("OWASP LLM Top 10 tab created successfully.")
print("Compliance Gaps tab created successfully.")
print()
print("OUTPUT FILE:", output_file)
print("STATUS: SUCCESS")
print("=" * 65)